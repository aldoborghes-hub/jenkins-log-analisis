import re
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def format_duration(ms):
    if pd.isnull(ms):
        return ""
    total_seconds = int(ms / 1000)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours}h {minutes}m {seconds}s"

def extraer_procesos(raw_lines):
    EXCLUDE_TAGS = ("CR", "PF", "RQ")
    VALID_NOMBRES = ("Orchestrator", "-", "EnvironmentPreparation")
    filas = []
    patron = re.compile(
        r"\[(\d{4}\.\d{2}\.\d{2} \d{2}:\d{2}:\d{2})\]\s+\[([^\]]+)\]\s+\[([^\]]+)\]\s+\[([^\]]+)\]\s+(?:La fase|The) (\w+) (?:ha finalizado|phase has ended) in (\d+)ms",
        re.IGNORECASE
    )
    for line in raw_lines:
        m = patron.search(line)
        if m:
            fecha_str, nombre, tipo, descripcion, fase, duracion_ms = m.groups()
            if nombre not in VALID_NOMBRES:
                continue
            if any(tag in descripcion for tag in EXCLUDE_TAGS):
                continue
            duracion_ms = int(duracion_ms)
            if duracion_ms == 0:
                continue
            filas.append({
                "Nombre": nombre,
                "Tipo": tipo,
                "Descripcion": descripcion,
                "Fase": fase.upper(),
                "Inicio": "",
                "Fin": fecha_str,
                "Duracion (ms)": duracion_ms,
                "Duracion (h:m:s)": format_duration(duracion_ms)
            })
    columnas = ["Nombre", "Tipo", "Descripcion", "Fase", "Inicio", "Fin", "Duracion (ms)", "Duracion (h:m:s)"]
    return pd.DataFrame(filas, columns=columnas)


def detectar_errores_por_etiqueta(raw_lines):
    import re
    errores = {}
    current_tag = None
    capture_next = False
    for line in raw_lines:
        tag_match = re.search(r"\[(CR-\d+)\]", line)
        if tag_match:
            current_tag = tag_match.group(1)
        if "The following error occurred while executing this line:" in line:
            capture_next = True
            continue
        if capture_next and current_tag:
            siguiente = line.strip().strip('"')
            if siguiente:
                errores[current_tag] = siguiente
            capture_next = False
    return errores

def generar_metricas(log_path_str):
    log_path = Path(log_path_str)

    # Detectar y corregir nombres mal formateados
    if log_path.name.endswith("_l.html_tiempos"):
        nuevo_nombre = log_path.name.replace("_l.html_tiempos", ".html")
        nuevo_path = log_path.with_name(nuevo_nombre)
        log_path.rename(nuevo_path)
        print(f"✔ Archivo renombrado internamente a: {nuevo_path.name}")
        log_path = nuevo_path  # actualizar la referencia para el resto del proceso

    with open(log_path, "r", encoding="utf-8") as file:
        raw_lines = file.readlines()

    patron = re.compile(
        r"\[(\d{4}\.\d{2}\.\d{2} \d{2}:\d{2}:\d{2})\]\s+\[(\w+)\].*?\[(CR|PF|RQ)-(\d+)\].*?The (\w+)\sphase has\s(ended|started)",
        re.IGNORECASE
    )

    eventos = []
    for line in raw_lines:
        match = patron.search(line)
        if match:
            fecha_str, tecnologia, tipo_etq, num_etq, fase, evento = match.groups()
            fecha = datetime.strptime(fecha_str.replace(".", "-"), "%Y-%m-%d %H:%M:%S")
            eventos.append({
                "fecha": fecha,
                "evento": evento.upper(),
                "etiqueta": f"{tipo_etq}-{num_etq}",
                "tecnologia": tecnologia,
                "fase": fase.upper()
            })

    df_eventos = pd.DataFrame(eventos)
    filas = []
    for (etiqueta, fase), grupo in df_eventos.groupby(["etiqueta", "fase"]):
        fecha_ini = grupo[grupo["evento"] == "STARTED"]["fecha"].min()
        fecha_fin = grupo[grupo["evento"] == "ENDED"]["fecha"].max()
        tecnologias = grupo["tecnologia"].unique()
        if pd.notnull(fecha_ini) and pd.notnull(fecha_fin) and fecha_fin > fecha_ini:
            duracion_ms = int((fecha_fin - fecha_ini).total_seconds() * 1000)
            filas.append({
                "etiqueta": etiqueta,
                "tecnologia": ", ".join(tecnologias),
                "fase": fase,
                "inicio": fecha_ini,
                "fin": fecha_fin,
                "duracion_ms": duracion_ms,
                "duracion_hms": format_duration(duracion_ms)
            })

    df_global = pd.DataFrame(filas)
    df_tiempos = df_global.groupby("tecnologia").agg(
        inicio=("inicio", "min"),
        fin=("fin", "max"),
        duracion_ms=("duracion_ms", "sum")
    ).reset_index()
    df_tiempos["duracion_hms"] = df_tiempos["duracion_ms"].apply(format_duration)

    
    df_etiquetas = df_global.groupby("etiqueta").agg(
        tecnologia=("tecnologia", lambda x: ", ".join(sorted(set(sum((s.split(", ") for s in x), []))))),
        inicio=("inicio", "min"),
        fin=("fin", "max")
    ).reset_index()
    df_etiquetas["duracion_ms"] = (df_etiquetas["fin"] - df_etiquetas["inicio"]).dt.total_seconds() * 1000
    df_etiquetas["duracion_ms"] = df_etiquetas["duracion_ms"].astype(int)
    df_etiquetas["duracion_hms"] = df_etiquetas["duracion_ms"].apply(format_duration)
    
    errores_etiqueta = detectar_errores_por_etiqueta(raw_lines)
    df_etiquetas["comentarios"] = df_etiquetas["etiqueta"].map(errores_etiqueta).fillna("")

    df_medias = df_global.groupby("tecnologia").agg(
        numero_etiquetas=("etiqueta", "count"),
        suma_total_ms=("duracion_ms", "sum")
    ).reset_index()
    df_medias["media_ms"] = (df_medias["suma_total_ms"] / df_medias["numero_etiquetas"]).astype(int)
    df_medias["media_hms"] = df_medias["media_ms"].apply(format_duration)

    # ----- PROCESOS -----
    df_procesos = extraer_procesos(raw_lines)

    wb = Workbook()
    ws = wb.active
    ws.title = "Portada"
    ws["A1"] = "Autor:"
    ws["B1"] = "Proyecto Público"
    ws["A2"] = "Organización:"
    ws["B2"] = "—"
    ws["A4"] = "Descripción:"
    ws["B4"] = "Métricas generadas a partir de log HTML de promoción Jenkins"

    def add_sheet(wb, name, df):
        ws = wb.create_sheet(name)
        ws.append(df.columns.tolist())
        for row in df.itertuples(index=False):
            ws.append(row)

    if ws.title == "Etiquetas" and "comentarios" in df.columns:
        from openpyxl.styles import PatternFill
        fill_rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            comentario = row[df.columns.get_loc("comentarios")]
            if comentario.value:
                comentario.fill = fill_rojo
        for col_idx, _ in enumerate(df.columns, start=1):
            max_length = max((len(str(cell.value)) for cell in ws[get_column_letter(col_idx)] if cell.value), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 40)

    add_sheet(wb, "GlobalData", df_global)
    add_sheet(wb, "Tiempos", df_tiempos)
    add_sheet(wb, "Etiquetas", df_etiquetas)
    add_sheet(wb, "Medias Tecnologia", df_medias)
    add_sheet(wb, "Procesos", df_procesos)  # <<--- ¡SIEMPRE crea esta hoja!

    nombre_base = log_path.name[:45].replace(" ", "_").replace("#", "_")
    output_path = Path(log_path.parent) / f"{nombre_base}_METRICAS.xlsx"
    wb.save(output_path)
    print(f"✔ Errores detectados en etiquetas: {sum(1 for v in errores_etiqueta.values() if v)}")

    print(f"Excel generado: {output_path}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        generar_metricas(sys.argv[1])
    else:
        print("Uso: python -m logparser.generar_metricas ruta_al_log.html")
